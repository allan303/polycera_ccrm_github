#!/usr/bin/env python
# -*- encoding: utf-8 -*-
# Author : Jack Li
# Email  : allanth3@163.com
# Date   : 2021-12-08

'''
Summary: Excel工具包
'''
import io

from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from typing import Any
from .config import STATIC_DEST


def wb_to_xlsx_file(wb: Workbook,
                    to_file: bool = False,
                    filename: str = None,
                    ):
    if not isinstance(wb, Workbook):
        raise TypeError('WB必须是Workbook')
    filename = filename or 'wb输出'
    filename = filename + '.xlsx'
    if to_file:
        wb.save(filename)
        return None
    path2 = io.BytesIO()  # 内存中的位置
    wb.save(path2)  # 保存到内存中
    path2.flush()
    path2.seek(0)
    return path2


def pd_to_wb(pd_class: BaseModel, data: Any, title: str = None, description: str = None) -> Workbook:
    '''
    Summary: 通过pd写入data
    '''
    header_row_index = 1
    merge_row_index = 1
    # style

    fill = PatternFill(fill_type=None,
                       start_color='FFFFFFFF',
                       end_color='FF000000')
    border = Border(left=Side(border_style=None,
                              color='FF000000'),
                    right=Side(border_style=None,
                               color='FF000000'),
                    top=Side(border_style=None,
                             color='FF000000'),
                    bottom=Side(border_style=None,
                                color='FF000000'),
                    diagonal=Side(border_style=None,
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                 color='FF000000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )
    alignment = Alignment(horizontal='general',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)
    number_format = 'General'
    protection = Protection(locked=True,
                            hidden=False)
    # workbook
    wb = Workbook()
    ws = wb.active
    cols = pd_class._columns
    max_col_index = len(cols)
    # 标题
    if title:
        row = ws.append([title])
        ws.merge_cells(start_row=merge_row_index, start_column=1,
                       end_row=merge_row_index, end_column=max_col_index)
        ws[f'A{merge_row_index}'].font = Font(name='Calibri',
                                              size=12,
                                              bold=True,
                                              italic=False,
                                              vertAlign=None,
                                              underline='none',
                                              strike=False,
                                              color='074F97')
        header_row_index += 1
        merge_row_index += 1
    # 副标题
    if description:
        ws.append([description])
        ws.merge_cells(start_row=merge_row_index, start_column=1,
                       end_row=merge_row_index, end_column=max_col_index)
        ws[f'A{merge_row_index}'].font = Font(color='074F97')
        header_row_index += 1
        merge_row_index += 1
    # 表头
    ws.append(cols)

    # row_header = ws[]
    # cells.fill = PatternFill(fill_type='solid',
    #                          start_color='074F97')
    # cells.font = Font(color='FFFFFF')
    # 数据
    for x in data:
        pd = pd_class.from_orm(x)
        ws.append(list(pd.dict().values()))
    return wb
